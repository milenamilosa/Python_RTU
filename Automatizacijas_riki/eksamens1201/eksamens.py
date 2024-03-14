# #uzd2
# from openpyxl import Workbook, load_workbook 
# wb=load_workbook('data.xlsx')
# ws=wb['Lapa_4']
# max_row=ws.max_row
# s=[]
# sum = 0
# for row in range(2,max_row+1):
#     a=(ws['H' + str(row)].value)
#     if a == "Critical":
#         b = (ws['L' + str(row)].value)
#         if int(b)>=10 and int(b)<=50:
#             sum+=1
#     # s.append(a)
# print(sum)



# import pandas
# import openpyxl
# # get_info=input() #input information from terminal
# fails = pandas.read_excel("EmployeeData.xlsx", sheet_name="Data") # if no pages are specified, then the last one saved is open
# info_list = fails.values.tolist()
# print(info_list)
# count = 0
# for line in info_list:
#     if line[3] == "Female":
#         count +=1
# print(count)

# # uzd3
# from openpyxl import Workbook, load_workbook 
# wb=load_workbook('data.xlsx')
# ws=wb['Lapa_4']
# max_row=ws.max_row
# s=[]
# sum = 0
# for row in range(4,max_row+1):
#     a=(ws['D' + str(row)].value)
#     if a and a[0:2] == "Ab":
#         b = (ws['L' + str(row)].value)
#         if int(b)<10:
#             sum+=1
# print(sum)

# #uzd3
# from openpyxl import Workbook, load_workbook 
# wb=load_workbook('data.xlsx')
# ws=wb['Lapa_2']
# max_row=ws.max_row
# s=[]
# sum = 0
# for row in range(4,max_row+1):
#     a=(ws['H' + str(row)].value)
#     if a == "Lane Cove" or a == "Greenwich":
#         b = (ws['L' + str(row)].value)
#         if "1" in b:
#             sum+=1
# print(sum)

# #uzd4 ??????????????????
# from openpyxl import Workbook, load_workbook 
# wb=load_workbook('data.xlsx')
# ws=wb['Lapa_2']
# max_row=ws.max_row
# s=[]
# sum = 0
# for row in range(4,max_row+1):
#     a=(ws['H' + str(row)].value)
#     if a == "Lane Cove" or a == "Greenwich":
#         b = (ws['L' + str(row)].value)
#         if b == "1" or b == "10" or b == "11" or b == "12" or b == "13" or b == "14" or b == "15" or b == "16" or b == "17" or b == "18" or b == "19" or b == "21":
#             sum+=1
#         if b == "11":
#             sum+=2
# print(sum)

# #uzd5
# from openpyxl import Workbook, load_workbook 
# wb=load_workbook('data.xlsx')
# ws=wb['Lapa_3']
# max_row=ws.max_row
# s=[]
# sum = 0
# for row in range(4,max_row+1):
#     a=(ws['L' + str(row)].value)
#     if "AVENUE" in a:
#         sum+=1
# print(sum)

# #uzd6
# from openpyxl import Workbook, load_workbook 
# wb=load_workbook('data.xlsx')
# ws=wb['Lapa_3']
# max_row=ws.max_row
# s=[]
# sum = 0
# for row in range(4,max_row+1):
#     a=(ws['F' + str(row)].value)
#     if a[0:2] == "W3":
#         b=(ws['B' + str(row)].value)
#         sum+=float(b)
# print(sum)

#uzd7
from openpyxl import Workbook, load_workbook 
wb=load_workbook('data.xlsx')
ws1=wb['Lapa_1']
ws2=wb['Lapa_2']
max_row1=ws1.max_row
max_row2=ws2.max_row
s=[]
sum = 0
listA = []
listB = []

for row in range(2,max_row1+1):
    a=(ws1['A' + str(row)].value)
    listA.append(a)
print(listA)

for row in range(4,max_row2+1):
    b=(ws2['A' + str(row)].value)
    listB.append(b)
print(listB)

# Noskaidrojiet, cik daudz (skaits) ieraksti no 2. saraksta atkārtojas pirmajā sarakstā?  (33)
# Noskaidrojiet, cik daudz (skaits) ieraksti no 1. saraksta atkārtojas otrajā sarakstā?  (39)
for element in listB:
    if element in listA:
        sum+=1
print(sum)