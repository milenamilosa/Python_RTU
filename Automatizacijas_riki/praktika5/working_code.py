import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)


file_path = 'C:/Users/milos/Downloads/people.csv' # C:\Users\milos\Desktop\imp\Uni\VSC\Python_RTU\praktika5\salary.xlsx
data = pd.read_csv(file_path)
name = (data['First Name'] + ' ' + data['Last Name']).tolist()


def get_encoded_name(full_name):
    driver.get("https://emn178.github.io/online-tools/crc32.html")
   
    find = driver.find_element(By.ID, "input")
    find.send_keys(full_name)
    output = driver.find_element(By.ID, "output").get_attribute("value")
    find.clear()
    return output


encoded_names = {}
for employee_name in name:
    encoded_name = get_encoded_name(employee_name)
    encoded_names[employee_name] = encoded_name


workbook = load_workbook(filename='C:/Users/milos/Downloads/salary.xlsx')
sheet = workbook.active


salaries = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
    encoded_name = row[0]
    salary = row[1]
    for employee, encoded in encoded_names.items():
        if encoded == encoded_name:
            if employee not in salaries:
                salaries[employee] = [salary]
            else:
                salaries[employee].append(salary)


for employee, salary_list in salaries.items():
    total_salary = sum(salary_list)
    print(f"Employee: {employee}, Total Salary: {total_salary}")


driver.quit()