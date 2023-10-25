from openpyxl import Workbook, load_workbook 
wb=load_workbook('C:\Users\milos\Desktop\Uni\VSC\Python\praktika1\test1.xlsx')
ws=wb.active
total=0
temp=ws['B2'].value
print(temp)

print(total)
wb.close()